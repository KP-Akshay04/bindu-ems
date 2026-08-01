from flask import Blueprint, request, jsonify
from openpyxl import load_workbook
import tempfile
import os

import_bp = Blueprint("import_bp", __name__)

REQUIRED_COLUMNS = [
    "ID NO",
    "NAME",
    "LOCATION",
    "department",
    "designation"
]


@import_bp.route("/api/import/validate", methods=["POST"])
def validate_import():

    if "file" not in request.files:
        return jsonify({
            "success": False,
            "message": "No file uploaded."
        }), 400

    file = request.files["file"]

    if file.filename == "":
        return jsonify({
            "success": False,
            "message": "Please select an Excel file."
        }), 400

    temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    file.save(temp.name)

    try:

        workbook = load_workbook(temp.name)
        sheet = workbook.active

        headers = []

        for cell in sheet[1]:
            headers.append(str(cell.value).strip() if cell.value else "")

        missing = [
            col for col in REQUIRED_COLUMNS
            if col not in headers
        ]

        if missing:
            return jsonify({
                "success": False,
                "message": "Invalid template.",
                "missing_columns": missing
            }), 400

        total_rows = sheet.max_row - 1

        preview = []

        for row in sheet.iter_rows(min_row=2, max_row=11, values_only=True):

            preview.append({
                "employee_code": row[0],
                "name": row[1],
                "location": row[2],
                "department": row[3],
                "designation": row[4]
            })

        return jsonify({

            "success": True,

            "message": "Excel validated successfully.",

            "total_rows": total_rows,

            "preview": preview

        })

    except Exception as e:

        return jsonify({
            "success": False,
            "message": str(e)
        }), 500

    finally:

        if os.path.exists(temp.name):
            os.remove(temp.name)