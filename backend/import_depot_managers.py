from openpyxl import load_workbook

from app import create_app, db
from app.models.employee import Employee
from app.models.branch import Branch
from app.models.department import Department
from app.models.designation import Designation
from app.utils.security import hash_password

EXCEL_FILE = "../Depo EMP List.xlsx"
DEFAULT_PASSWORD = "bindu@123"


def normalize(text):
    return (
        str(text)
        .strip()
        .lower()
        .replace("  ", " ")
    )


app = create_app()

with app.app_context():

    print("=" * 60)
    print("IMPORTING DEPOT MANAGERS")
    print("=" * 60)

    workbook = load_workbook(EXCEL_FILE)

    sheet = workbook.active

    # Row 2 contains headers
    headers = [str(c.value).strip() for c in sheet[2]]

    id_col = headers.index("ID NO")
    name_col = headers.index("NAME")
    location_col = headers.index("LOCATION")
    department_col = headers.index("department")
    designation_col = headers.index("designation")

    depot_department = Department.query.filter(
        Department.department_name.ilike("Depot")
    ).first()

    depot_designation = Designation.query.filter(
        Designation.designation_name.ilike("Depot Executive")
    ).first()

    if not depot_department:
        raise Exception("Depot department not found.")

    if not depot_designation:
        raise Exception("Depot Executive designation not found.")

    imported = 0
    skipped = 0
    failed = 0

    for row in sheet.iter_rows(min_row=3, values_only=True):

        if row[0] is None:
            continue

        employee_code = str(row[id_col]).strip()

        full_name = str(row[name_col]).strip()

        location = str(row[location_col]).strip()

        existing = Employee.query.filter_by(
            employee_code=employee_code
        ).first()

        if existing:
            print(f"SKIPPED : {employee_code}")
            skipped += 1
            continue

        BRANCH_ALIASES = {
            "LACKNOW": "LUCKNOW",
            "GOREGAV": "GOREGAON",
            "CHENAI": "CHENNAI",
            "ERANAKUALM": "ERNAKULAM",
            "ERANKULAM": "ERNAKULAM",
        }

        location = BRANCH_ALIASES.get(
        location.upper(),
        location
        )

        branch = Branch.query.filter(
            Branch.branch_name.ilike(location)
        ).first()

        if not branch:
            print(f"BRANCH NOT FOUND : {location}")
            failed += 1
            continue

        email = (
            employee_code
            .replace(" ", "")
            .lower()
            + "@binduems.local"
        )

        employee = Employee(

            employee_code=employee_code,

            full_name=full_name,

            email=email,

            phone="",

            password_hash=hash_password(DEFAULT_PASSWORD),

            branch_id=branch.branch_id,

            department_id=depot_department.department_id,

            designation_id=depot_designation.designation_id,

            role="Employee",

            status="Active",

            basic_salary=0,

            leave_balance=12

        )

        db.session.add(employee)

        imported += 1

    db.session.commit()

    print()
    print("=" * 60)
    print("IMPORT FINISHED")
    print("=" * 60)

    print("Imported :", imported)
    print("Skipped  :", skipped)
    print("Failed   :", failed)